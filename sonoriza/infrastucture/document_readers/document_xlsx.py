from pathlib import Path
from openpyxl import load_workbook
from sonoriza.core.services.document_reader import IDocumentReader

class XLSXReader(IDocumentReader):
    @property
    def supported_formats(self) -> list[str]:
        return ['xlsx', 'xls']

    def extract_text(self, file_path: Path) -> str:
        """Extrae texto de archivos Excel, concatenando columnas por fila."""
        workbook = load_workbook(filename=file_path, read_only=True)
        full_text = []
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                # Filtra celdas no vacías y convierte a string
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():  # Omite filas vacías
                    full_text.append(row_text)
        
        return '\n'.join(full_text)